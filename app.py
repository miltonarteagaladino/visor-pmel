import os
import re
import textwrap
import itertools
from collections import Counter, defaultdict
import pandas as pd
import streamlit as st
import openpyxl
import plotly.graph_objects as go
import plotly.express as px
from pyvis.network import Network
import streamlit.components.v1 as components

# BLOQUEO ABSOLUTO DE PYARROW (ANTI-CRASH PARA MAC)
os.environ["ARROW_USER_SIMD_LEVEL"] = "NONE"
os.environ["STREAMLIT_SERVER_MAX_MESSAGE_SIZE"] = "200"

st.set_page_config(layout="wide", page_title="PMEL - Fundación Corona", initial_sidebar_state="expanded")

# --- 1. IDENTIDAD CORPORATIVA Y NAVEGACIÓN ---
st.markdown("""
    <style>
    .stApp { background-color: #F8F9FA; }
    .corona-header { background-color: #003366 !important; padding: 20px; border-radius: 10px; text-align: center; margin-bottom: 25px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
    .corona-title { color: #FFFFFF !important; margin: 0; font-family: 'Helvetica Neue', sans-serif; font-size: 32px; font-weight: 700; }
    .corona-subtitle { color: #FFB300 !important; margin: 5px 0 0 0; font-family: 'Helvetica Neue', sans-serif; font-size: 16px; }
    </style>
""", unsafe_allow_html=True)

with st.sidebar:
    try:
        st.image("logo.png", use_container_width=True)
    except:
        pass
        
    st.markdown("### Menú de Navegación")
    pagina_actual = st.radio("Ir a:", [
        "🕸️ Mapa Sistémico (Redes)", 
        "📊 Analítica de Portafolio", 
        "🧬 Patrones de Co-Ocurrencia",
        "📈 Reporte Ejecutivo MEL",
        "📊 Indicadores Ejecutivos para Dirección"
    ])
    st.markdown("---")

titulo_limpio = pagina_actual.split(" ", 1)[1] if " " in pagina_actual else pagina_actual
st.markdown(f"""
    <div class="corona-header">
        <h1 class="corona-title">{titulo_limpio}</h1>
        <p class="corona-subtitle">ESTRATEGIA 2030 | Fundación Corona</p>
    </div>
""", unsafe_allow_html=True)

# --- 2. DECODIFICADORES GLOBALES ---
MAPA_AREAS = {
    'ED': 'Educación', 'EM': 'Empleo', 'LE': 'Libre Elección', 'IN': 'Incidencia',
    'AA': 'Aprendizaje/Adaptación', 'TC': 'Transformación Cultural', 'CC': 'Comunicaciones',
    'DE': 'Dirección Ejecutiva', 'SA': 'Servicios Administrativos'
}

def extraer_area_codigo(codigo):
    partes = codigo.split('-')
    for p in partes:
        if p in MAPA_AREAS: return f"{p} - {MAPA_AREAS[p]}"
    for clave, valor in MAPA_AREAS.items():
        if f"-{clave}-" in codigo or f"-{clave}" in codigo: return f"{clave} - {valor}"
    return "Múltiples / Otra Área"

def detectar_accion_oficial(texto):
    t = str(texto).lower()
    if "analiza" in t or "comprend" in t: return "1. Analizamos y comprendemos el contexto"
    if "conocimiento" in t or "evidencia" in t: return "2. Generamos conocimiento y evidencia"
    if "agenda" in t: return "3. Posicionamos agendas"
    if "colectiv" in t: return "4. Promovemos acciones colectivas"
    if "diseño" in t or "solucion" in t: return "5. Promovemos el diseño y desarrollo de soluciones"
    if "capacidad" in t or "lider" in t: return "6. Fortalecemos capacidades (liderazgo)"
    return None

def obtener_color_borde_categoria(col_index):
    if col_index == 4: return "#9C27B0" 
    elif 5 <= col_index <= 10: return "#C0CA33" 
    elif 11 <= col_index <= 18: return "#D81B60" 
    elif 19 <= col_index <= 22: return "#FFB300" 
    elif 23 <= col_index <= 26: return "#4CAF50" 
    return "#BDBDBD"

def limpiar_codigo_historia(codigo):
    c_limpio = str(codigo).strip().upper().replace(" ", "")
    c_limpio = c_limpio.replace(".", "-").replace("--", "-")
    if "02EM-FC" in c_limpio: c_limpio = c_limpio.replace("02EM-FC", "02-EM-FC")
    return c_limpio

def obtener_estado_por_sufijo(codigo):
    cod_upper = str(codigo).upper().strip()
    if cod_upper.endswith("-EC"): return 'Negro (Ejemplo de Cambio)'
    if cod_upper.endswith("-BC"): return 'Rojo (Señal de Buen Camino)'
    if cod_upper.endswith("-IC"): return 'Naranja (Intención de Cambio)'
    if cod_upper.endswith("-EE"): return 'Morado (Efecto Estancado)'
    return 'Negro (Ejemplo de Cambio)' 

def extraer_corazon_codigo(codigo):
    c_limpio = limpiar_codigo_historia(codigo)
    if c_limpio.endswith(("-EC", "-IC", "-EE", "-BC")):
        c_limpio = c_limpio[:-3]
    match = re.search(r'([A-Z]+-\w+-\d+-[A-Z]+-[A-Z]+)', c_limpio)
    if match: return match.group(1)
    match2 = re.search(r'([A-Z]+-\d+)', c_limpio)
    if match2: return match2.group(1)
    return c_limpio

def acortar_codigo(codigo):
    corazon = extraer_corazon_codigo(codigo)
    partes = corazon.split('-')
    if len(partes) >= 3: return f"{partes[1]}{partes[2]}"
    return corazon

def formatear_caja(texto, ancho=35):
    lineas = textwrap.wrap(texto, width=ancho)
    if len(lineas) > 3: return "\n".join(lineas[:3]) + "..."
    return "\n".join(lineas)

# --- 3. EXTRACCIÓN ESTRUCTURAL JERÁRQUICA (CATÁLOGO MAESTRO) ---
@st.cache_data
def extraer_universo_y_conexiones(ruta_archivo):
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    
    catalogo_iniciativas = []
    map_portafolio = {}
    map_nombre = {}
    
    if "Matriz Portafolio" in wb.sheetnames:
        ws_port = wb["Matriz Portafolio"]
        header_row, col_acro, col_port, col_nom = -1, -1, -1, -1
        
        for r in range(1, 10):
            for c in range(1, ws_port.max_column + 1):
                val = str(ws_port.cell(row=r, column=c).value).strip()
                if val == "Nombre": col_acro = c; header_row = r
                elif val == "Oportunidad a la que más aporta": col_port = c
                elif val == "Mecanismo de acción": col_nom = c
            if col_acro != -1: break
                
        if col_acro != -1 and col_port != -1:
            for r in range(header_row + 1, ws_port.max_row + 1):
                acronimo = str(ws_port.cell(row=r, column=col_acro).value).strip()
                porta = str(ws_port.cell(row=r, column=col_port).value).strip()
                nom = str(ws_port.cell(row=r, column=col_nom).value).strip() if col_nom != -1 else acronimo
                if acronimo and acronimo not in ['None', 'nan', '']:
                    porta_limpio = porta if porta and porta != 'nan' else 'Sin Portafolio Asignado'
                    catalogo_iniciativas.append({
                        'Acrónimo': acronimo,
                        'Iniciativa': nom,
                        'Portafolio': porta_limpio
                    })
                    map_portafolio[acronimo] = porta_limpio
                    map_nombre[acronimo] = nom

    def get_portafolio_y_nombre(sheet_name):
        if sheet_name in map_portafolio: return map_portafolio[sheet_name], map_nombre.get(sheet_name, sheet_name)
        if '-' in sheet_name:
            sin_prefijo = sheet_name.split('-', 1)[1]
            if sin_prefijo in map_portafolio: return map_portafolio[sin_prefijo], map_nombre.get(sin_prefijo, sheet_name)
        return "Sin Portafolio Asignado", sheet_name

    datos = []
    pestañas = [sht for sht in wb.sheetnames if sht.startswith('I-') or sht.startswith('M-') or sht.startswith('B-') or sht.startswith('H-')]
    
    for sheet_name in pestañas:
        ws = wb[sheet_name]
        porta_asignado, iniciativa_asignada = get_portafolio_y_nombre(sheet_name)
        
        fila_cabeceras = -1
        for r in range(1, min(30, ws.max_row + 1)):
            val = ws.cell(row=r, column=1).value
            if val and str(val).strip().lower().startswith("acciones est"):
                fila_cabeceras = r
                break
        if fila_cabeceras == -1: continue
            
        colores_cambios_font = {}
        for col in range(3, min(18, ws.max_column + 1)):
            val_cambio = ws.cell(row=fila_cabeceras, column=col).value
            if val_cambio and str(val_cambio).strip() not in ['None', '']:
                c_texto = re.sub(r'\s+', ' ', str(val_cambio).replace('\n', ' ').strip())
                colores_cambios_font[c_texto] = obtener_color_borde_categoria(col)

        textos_hist = {}
        for row in range(1, fila_cabeceras):
            cod = ws.cell(row=row, column=1).value
            txt = ws.cell(row=row, column=2).value
            if cod and isinstance(cod, str) and '-' in cod:
                c_limpio = extraer_corazon_codigo(cod)
                textos_hist[c_limpio] = str(txt).strip() if txt else "Sin narrativa documentada."

        accion_actual = None
        for row in range(fila_cabeceras + 1, ws.max_row + 1):
            val_accion = ws.cell(row=row, column=1).value
            val_str = str(val_accion).strip() if val_accion else ""
            if val_str.lower().startswith('antecedentes') or val_str.lower().startswith('contexto'): break 
                
            nueva_accion = detectar_accion_oficial(val_str)
            if nueva_accion: accion_actual = nueva_accion
            if not accion_actual: continue 
            
            for col in range(3, min(18, ws.max_column + 1)):
                val_cambio = ws.cell(row=fila_cabeceras, column=col).value
                if not val_cambio or str(val_cambio).strip() in ['None', '']: continue
                
                cambio_texto = re.sub(r'\s+', ' ', str(val_cambio).replace('\n', ' ').strip())
                
                val_conexion = ws.cell(row=row, column=col).value
                if val_conexion and isinstance(val_conexion, str) and '-' in val_conexion:
                    codigos = [cd.strip() for cd in re.split(r'[\n,;\s]+', val_conexion) if cd.strip() and '-' in cd]
                    for codigo in codigos:
                        cod_limpio = limpiar_codigo_historia(codigo)
                        c_corto = extraer_corazon_codigo(cod_limpio)
                        estado_nom = obtener_estado_por_sufijo(cod_limpio)
                        color_fuente = colores_cambios_font.get(cambio_texto, "#BDBDBD")
                        
                        datos.append({
                            'Portafolio': porta_asignado,
                            'Iniciativa': iniciativa_asignada, 
                            'Acción Estratégica': accion_actual,
                            'Cambio Esperado': cambio_texto, 
                            'Historia_Cod': c_corto, 
                            'Historia_Corta': acortar_codigo(c_corto),
                            'Texto': textos_hist.get(c_corto, "Narrativa no encontrada."), 
                            'Estado': estado_nom,
                            'Color_Borde': color_fuente,
                            'Área': extraer_area_codigo(c_corto)
                        })
                        
    return catalogo_iniciativas, datos

catalogo_iniciativas, datos_list = extraer_universo_y_conexiones("Matriz.xlsx")
df_catalogo = pd.DataFrame(catalogo_iniciativas)
df_conexiones = pd.DataFrame(datos_list)

# --- DIAGNÓSTICO ESTRICTO EN BARRA LATERAL ---
with st.sidebar:
    st.markdown("### 🛠️ Diagnóstico Estructural")
    st.write(f"**Universo Iniciativas:** {len(df_catalogo)}")
    st.write(f"**Registros (Conexiones):** {len(datos_list)}")
    st.write(f"**Historias Únicas:** {len(set(d['Historia_Cod'] for d in datos_list))}")
    st.write(f"**Acciones Estratégicas:** {len(set(d['Acción Estratégica'] for d in datos_list))}")
    st.write(f"**Cambios Reales:** {len(set(d['Cambio Esperado'] for d in datos_list))}")
    st.markdown("---")

acciones_unicas = sorted(list(set(d['Acción Estratégica'] for d in datos_list)))
cambios_unicos = sorted(list(set(d['Cambio Esperado'] for d in datos_list)))
dict_acciones = {acc: f"A{i+1}" for i, acc in enumerate(acciones_unicas)}
dict_cambios = {cam: f"C{i+1}" for i, cam in enumerate(cambios_unicos)}

# COLORES GENERALES 
COLORES_ESTADO = {'Negro (Ejemplo de Cambio)': '#212121', 'Naranja (Intención de Cambio)': '#FF9800', 'Rojo (Señal de Buen Camino)': '#D32F2F', 'Morado (Efecto Estancado)': '#351C75'}
COLORES_HEX_PUROS = {'Negro (Ejemplo de Cambio)': '#212121', 'Naranja (Intención de Cambio)': '#FF9800', 'Rojo (Señal de Buen Camino)': '#D32F2F', 'Morado (Efecto Estancado)': '#7B1FA2'}

# COLORES EXCLUSIVOS PARA REPORTE MEL Y DIRECTIVO (Ajuste Visual)
MEL_COLORS = {
    'Negro (Ejemplo de Cambio)': '#4CAF50', # Verde
    'Rojo (Señal de Buen Camino)': '#FF9800', # Naranja
    'Naranja (Intención de Cambio)': '#2196F3', # Azul
    'Morado (Efecto Estancado)': '#F44336' # Rojo
}

# ==========================================
# GESTIÓN GLOBAL DE FILTROS (Basado en df_catalogo)
# ==========================================
if pagina_actual not in ["📈 Reporte Ejecutivo MEL", "📊 Indicadores Ejecutivos para Dirección"]:
    st.markdown("### 🎛️ Filtros Globales (Jerarquía Estricta)")
    col_fg1, col_fg2 = st.columns(2)
    with col_fg1:
        portafolios_disp = sorted([p for p in df_catalogo['Portafolio'].unique() if p != ''])
        port_sel = st.selectbox("🎯 1. Portafolio:", ["Todos los Portafolios"] + portafolios_disp)
        
        cat_f1 = df_catalogo if port_sel == "Todos los Portafolios" else df_catalogo[df_catalogo['Portafolio'] == port_sel]
        datos_f1 = datos_list if port_sel == "Todos los Portafolios" else [d for d in datos_list if d.get('Portafolio') == port_sel]
        
    with col_fg2:
        inis_disp = sorted(cat_f1['Iniciativa'].unique())
        ini_sel = st.selectbox("🌍 2. Iniciativa:", ["Todo el Portafolio"] + inis_disp)
        
        cat_ana = cat_f1 if ini_sel == "Todo el Portafolio" else cat_f1[cat_f1['Iniciativa'] == ini_sel]
        datos_ana = datos_f1 if ini_sel == "Todo el Portafolio" else [d for d in datos_f1 if d.get('Iniciativa') == ini_sel]
    st.markdown("---")

# ==========================================
# PÁGINA 1: MAPA SISTÉMICO 
# ==========================================
if pagina_actual == "🕸️ Mapa Sistémico (Redes)":
    
    col_f3, col_f4 = st.columns(2)
    with col_f3:
        est_sel = st.selectbox("🎨 3. Estado:", ["Todos los estados"] + sorted(list(set(d['Estado'] for d in datos_ana))))
        datos_f2 = datos_ana if est_sel == "Todos los estados" else [d for d in datos_ana if d['Estado'] == est_sel]
    with col_f4:
        hist_sel = st.selectbox("🔍 4. Historia Específica:", ["Todas las historias"] + sorted(list(set(f"{d['Historia_Corta']} ({d['Historia_Cod']})" for d in datos_f2))))
        df_final = datos_f2 if hist_sel == "Todas las historias" else [d for d in datos_f2 if d['Historia_Cod'] == hist_sel.split("(")[1].replace(")", "")]

    es_macronivel = (ini_sel == "Todo el Portafolio")
    col_ctrl1, col_ctrl2 = st.columns(2)
    with col_ctrl1: vista_simplificada = st.toggle("👁️ **Vista Simplificada** (Ocultar historias)", value=es_macronivel)
    with col_ctrl2: congelar_mapa = st.toggle("🛑 **Congelar Mapa** (Fijar cajas)", value=es_macronivel)

    col_grafo, col_lector = st.columns([7, 3])
    with col_grafo:
        if not df_final:
            st.warning("No hay datos para esta combinación.")
        else:
            peso_acciones = Counter([d['Acción Estratégica'] for d in df_final])
            peso_cambios = Counter([d['Cambio Esperado'] for d in df_final])
            borde_cambios_map = {d['Cambio Esperado']: d['Color_Borde'] for d in df_final}
            
            min_peso = 1
            max_peso = max(list(peso_acciones.values()) + list(peso_cambios.values()) + [1])
            def calc_fs(peso):
                if max_peso == min_peso: return 16
                return int(16 + ((peso - min_peso) / (max_peso - min_peso)) * 24)

            net = Network(height='700px', width='100%', directed=True, bgcolor='#FFFFFF', font_color='#202124', cdn_resources='remote')
            net.set_options(f"""
            var options = {{ "nodes": {{ "margin": 12, "borderWidth": 4, "borderWidthSelected": 6 }}, "edges": {{ "smooth": {{ "type": "dynamic" }}, "width": 2.5 }}, "layout": {{ "hierarchical": {{ "enabled": true, "direction": "LR", "levelSeparation": {600 if vista_simplificada else 380}, "nodeSpacing": 120 }} }}, "physics": {{ "enabled": {"false" if congelar_mapa else "true"}, "solver": "hierarchicalRepulsion" }} }}
            """)

            for acc in set(d['Acción Estratégica'] for d in df_final):
                fs = calc_fs(peso_acciones[acc])
                net.add_node(acc, label=formatear_caja(acc, 35), shape='box', level=1 if vista_simplificada else 2, color={'border': '#003366', 'background': '#E3F2FD'}, font={'color': '#003366', 'face': 'sans-serif', 'bold': True, 'size': fs})
            
            for cam in set(d['Cambio Esperado'] for d in df_final):
                fs = calc_fs(peso_cambios[cam])
                color_letra_excel = borde_cambios_map.get(cam, '#BDBDBD')
                net.add_node(cam, label=formatear_caja(cam, 35), shape='box', level=2 if vista_simplificada else 3, color={'background': '#FFFFFF', 'border': color_letra_excel}, font={'color': '#212121', 'face': 'sans-serif', 'bold': True, 'size': fs})

            if vista_simplificada:
                rutas = defaultdict(list)
                for d in df_final: rutas[(d['Acción Estratégica'], d['Cambio Esperado'], COLORES_HEX_PUROS.get(d['Estado'], '#000'), d['Estado'])].append(d['Historia_Corta'])
                track = Counter()
                for (acc, cam, color, estado), h_list in rutas.items():
                    track[(acc, cam)] += 1
                    idx = track[(acc, cam)]
                    curva = {"type": "straight"} if idx == 1 else {"type": "curvedCW", "roundness": 0.15 * (idx // 2) * (1 if idx % 2 == 0 else -1)}
                    net.add_edge(acc, cam, color=color, width=1.5 + (len(h_list)*1.5), title=f"Estado: {estado}\nHistorias: {', '.join(h_list)}", smooth=curva)
            else:
                historias_unicas = set(d['Historia_Cod'] for d in df_final)
                peso_historias = Counter([d['Historia_Cod'] for d in df_final])
                for hist in historias_unicas:
                    hist_corta = [d['Historia_Corta'] for d in df_final if d['Historia_Cod'] == hist][0]
                    fs_h = calc_fs(peso_historias[hist])
                    net.add_node(hist, label=hist_corta, color={'background': '#FFFFFF', 'border': '#E0E0E0'}, shape='ellipse', level=1, font={'size': fs_h})
                for d in df_final:
                    net.add_edge(d['Historia_Cod'], d['Acción Estratégica'], color='#E0E0E0')
                    net.add_edge(d['Acción Estratégica'], d['Cambio Esperado'], color=COLORES_HEX_PUROS.get(d['Estado'], '#000'), title=f"{d['Estado']}")

            html_source = net.generate_html()
            components.html(html_source, height=720)

    with col_lector:
        st.header("📖 Lector Narrativo")
        h_leer = st.selectbox("📚 Seleccionar historia:", ["Seleccionar..."] + sorted(list(set(f"{d['Historia_Corta']} ({d['Historia_Cod']})" for d in df_final))))
        if h_leer != "Seleccionar...":
            cod_real = h_leer.split("(")[1].replace(")", "")
            info_h = [d for d in datos_list if d['Historia_Cod'] == cod_real]
            if info_h:
                st.success(f"**Historia:** {info_h[0]['Historia_Corta']}")
                st.write(info_h[0]['Texto'])
                st.markdown("---")
                for d in info_h: st.markdown(f"- ➔ {d['Cambio Esperado']} (**{d['Estado']}**)")

# ==========================================
# PÁGINA 2 Y 3: LÓGICA DE ANALÍTICA EXISTENTE
# ==========================================
elif pagina_actual == "📊 Analítica de Portafolio":
    
    st.markdown("#### 📊 Métricas de Validación del Portafolio")
    num_ini = len(cat_ana) 
    num_hist = len(set(d['Historia_Cod'] for d in datos_ana))
    num_acc = len(set(d['Acción Estratégica'] for d in datos_ana))
    num_cam = len(set(d['Cambio Esperado'] for d in datos_ana))
    
    km1, km2, km3, km4 = st.columns(4)
    km1.metric("Iniciativas del Universo", num_ini)
    km2.metric("Historias Trazadas", num_hist)
    km3.metric("Nodos Únicos", f"{num_acc + num_cam}", f"{num_acc} Acciones | {num_cam} Cambios", delta_color="off")
    km4.metric("Conexiones Totales", len(datos_ana))
    
    if datos_ana:
        conteo_est = Counter([d['Estado'] for d in datos_ana])
        ke1, ke2, ke3, ke4 = st.columns(4)
        ke1.metric("⚫ Ejemplos de Cambio", conteo_est.get('Negro (Ejemplo de Cambio)', 0))
        ke2.metric("🔴 Señales Buen Camino", conteo_est.get('Rojo (Señal de Buen Camino)', 0))
        ke3.metric("🟠 Intenciones de Cambio", conteo_est.get('Naranja (Intención de Cambio)', 0))
        ke4.metric("🟣 Efectos Estancados", conteo_est.get('Morado (Efecto Estancado)', 0))
    st.markdown("---")

    if not datos_ana:
        st.warning("La iniciativa o portafolio seleccionado tiene 0 historias y 0 conexiones reportadas en la matriz.")
    else:
        def crear_grafico_ranking(lista_datos, key_obj, color_scale, titulo_eje):
            lista_elementos = [d[key_obj] for d in lista_datos]
            if not lista_elementos: return None
            conteo = Counter(lista_elementos)
            items = sorted(conteo.items(), key=lambda x: x[1], reverse=False)
            y_labels = ["<br>".join(textwrap.wrap(k, width=50)) for k, v in items]
            x_vals = [v for k, v in items]
            textos_completos = [k for k, v in items]
            fig = go.Figure(go.Bar(x=x_vals, y=y_labels, orientation='h', marker=dict(color=x_vals, colorscale=color_scale), customdata=textos_completos, hovertemplate="<b>%{customdata}</b><br>Frecuencia: %{x}<extra></extra>"))
            fig.update_layout(xaxis_title=titulo_eje, yaxis_title="", margin=dict(l=0, r=0, t=0, b=0), height=max(300, len(items)*40), template="plotly_white")
            return fig

        st.markdown("### 🏆 Enfoque Estratégico (Rankings Cruzados)")
        c_gen1, c_gen2 = st.columns(2)
        with c_gen1:
            fig1 = crear_grafico_ranking(datos_ana, 'Acción Estratégica', 'Blues', 'Número de Historias')
            if fig1: st.plotly_chart(fig1, use_container_width=True)
        with c_gen2:
            fig2 = crear_grafico_ranking(datos_ana, 'Cambio Esperado', 'Teal', 'Número de Historias')
            if fig2: st.plotly_chart(fig2, use_container_width=True)

        st.markdown("#### Logros en Terreno (Ejemplos + Señales de Buen Camino)")
        datos_logros = [d for d in datos_ana if d['Estado'] in ['Negro (Ejemplo de Cambio)', 'Rojo (Señal de Buen Camino)']]
        c_log1, c_log2 = st.columns(2)
        with c_log1:
            fig3 = crear_grafico_ranking(datos_logros, 'Acción Estratégica', 'Greens', 'Historias (Negras/Rojas)')
            if fig3: st.plotly_chart(fig3, use_container_width=True)
        with c_log2:
            fig4 = crear_grafico_ranking(datos_logros, 'Cambio Esperado', 'Greens', 'Historias (Negras/Rojas)')
            if fig4: st.plotly_chart(fig4, use_container_width=True)

        st.markdown("#### Intenciones a Futuro (Solo Naranjas)")
        datos_intenciones = [d for d in datos_ana if d['Estado'] == 'Naranja (Intención de Cambio)']
        c_int1, c_int2 = st.columns(2)
        with c_int1:
            fig5 = crear_grafico_ranking(datos_intenciones, 'Acción Estratégica', 'Oranges', 'Historias (Naranjas)')
            if fig5: st.plotly_chart(fig5, use_container_width=True)
        with c_int2:
            fig6 = crear_grafico_ranking(datos_intenciones, 'Cambio Esperado', 'Oranges', 'Historias (Naranjas)')
            if fig6: st.plotly_chart(fig6, use_container_width=True)

        st.markdown("---")
        st.markdown("### 🌡️ Termómetros de Eficacia (Las 7 Tortas)")
        conteo_gen = Counter([d['Estado'] for d in datos_ana])
        fig_gen = go.Figure(go.Pie(labels=list(conteo_gen.keys()), values=list(conteo_gen.values()), hole=0.4, marker=dict(colors=[COLORES_ESTADO.get(k, '#000') for k in conteo_gen.keys()])))
        fig_gen.update_layout(title_text="<b>PROMEDIO GENERAL DEL SISTEMA</b>", margin=dict(l=0, r=0, t=40, b=0), height=350, template="plotly_white")
        st.plotly_chart(fig_gen, use_container_width=True)

        cols_pie = st.columns(3)
        for i, accion in enumerate(acciones_unicas):
            datos_acc = [d['Estado'] for d in datos_ana if d['Acción Estratégica'] == accion]
            if datos_acc:
                conteo_acc = Counter(datos_acc)
                fig_acc = go.Figure(go.Pie(labels=list(conteo_acc.keys()), values=list(conteo_acc.values()), hole=0.5, marker=dict(colors=[COLORES_ESTADO.get(k, '#000') for k in conteo_acc.keys()])))
                fig_acc.update_layout(title_text=f"<span style='font-size:13px'><b>{accion[:40]}...</b></span>", showlegend=False, margin=dict(l=10, r=10, t=40, b=10), height=220, template="plotly_white")
                with cols_pie[i % 3]:
                    st.plotly_chart(fig_acc, use_container_width=True)

        st.markdown("---")
        st.markdown("### 🗺️ Densidad de Impacto: Acción vs Cambio")
        df_heat_temp = pd.DataFrame(datos_ana)
        df_heat_temp['Accion_Corta'] = df_heat_temp['Acción Estratégica'].map(dict_acciones)
        df_heat_temp['Cambio_Corto'] = df_heat_temp['Cambio Esperado'].map(dict_cambios)
        
        heat_df = pd.crosstab(df_heat_temp['Accion_Corta'], df_heat_temp['Cambio_Corto'])
        hover_text = []
        for accion_corta in heat_df.index:
            hover_row = []
            for cambio_corto in heat_df.columns:
                acc_real = [k for k, v in dict_acciones.items() if v == accion_corta][0]
                cam_real = [k for k, v in dict_cambios.items() if v == cambio_corto][0]
                cant = heat_df.loc[accion_corta, cambio_corto]
                hover_row.append(f"<b>Acción:</b> {acc_real}<br><b>Cambio:</b> {cam_real}<br><b>Conexiones:</b> {cant}")
            hover_text.append(hover_row)

        fig_heat = px.imshow(heat_df, color_continuous_scale='YlGnBu', text_auto=True, aspect="auto")
        fig_heat.update_traces(customdata=hover_text, hovertemplate="%{customdata}<extra></extra>")
        fig_heat.update_layout(margin=dict(l=0, r=0, t=10, b=0), xaxis_title="Cambios Esperados (C1, C2...)", yaxis_title="Acciones Estratégicas (A1, A2...)", template="plotly_white")
        st.plotly_chart(fig_heat, use_container_width=True)

elif pagina_actual == "🧬 Patrones de Co-Ocurrencia":
    st.info("💡 **Fórmulas de Éxito:** El algoritmo cruza EXCLUSIVAMENTE las conexiones que tienen evidencia de cambio (Negras) o son señales de buen camino (Rojas). Omitiendo intenciones futuras o bloqueos.", icon="🚀")
    datos_exitosos = [d for d in datos_list if d['Estado'] in ['Negro (Ejemplo de Cambio)', 'Rojo (Señal de Buen Camino)']]

    if not datos_exitosos:
        st.warning("El portafolio seleccionado no tiene conexiones exitosas.")
    else:
        def algoritmo_seguro(lista_dicts, key_agrupadora, key_objetivo, nombre_grupo, min_r, max_r):
            grupos = defaultdict(set)
            for d in lista_dicts:
                if d[key_objetivo]: grupos[d[key_agrupadora]].add(d[key_objetivo])
                
            nombres_grupos = list(grupos.keys())
            patrones_encontrados = set()
            
            for i in range(len(nombres_grupos)):
                for j in range(i + 1, len(nombres_grupos)):
                    interseccion = frozenset(grupos[nombres_grupos[i]].intersection(grupos[nombres_grupos[j]]))
                    if len(interseccion) >= min_r:
                        if len(interseccion) <= max_r: patrones_encontrados.add(interseccion)
                        if len(interseccion) <= 8:
                            for r in range(min_r, len(interseccion)):
                                for combo in itertools.combinations(interseccion, r):
                                    patrones_encontrados.add(frozenset(combo))

            resultados = {}
            for pat in patrones_encontrados:
                apariciones = [g for g, items in grupos.items() if pat.issubset(items)]
                if len(apariciones) > 1: resultados[pat] = apariciones
                    
            patrones_limpios = {}
            for pat, aps in resultados.items():
                es_redundante = False
                for pat_otro, aps_otro in resultados.items():
                    if pat != pat_otro and pat.issubset(pat_otro) and len(aps) == len(aps_otro):
                        es_redundante = True
                        break
                if not es_redundante:
                    patrones_limpios[tuple(sorted(list(pat)))] = aps
                    
            ordenados = sorted(patrones_limpios.items(), key=lambda x: (len(x[1]), len(x[0])), reverse=True)
            if not ordenados:
                st.write(f"No hay patrones exitosos recurrentes en {nombre_grupo}s.")
                return
            for combo, grupos_list in ordenados[:30]:
                freq = len(grupos_list)
                tamano = len(combo)
                lista_md = "\n".join([f"* {item}" for item in combo])
                donde_se_vio = ", ".join(sorted(grupos_list))
                st.success(f"**[Fórmula de {tamano} Nodos] - Funcionó en {freq} {nombre_grupo}s distintos:**\n\n{lista_md}", icon="🎯")
                with st.expander(f"👁️ Ver en cuáles {nombre_grupo}s funcionó esto"):
                    st.write(f"**{nombre_grupo}s:** {donde_se_vio}")

        def algoritmo_mixto_veloz(lista_dicts, min_r=2, max_r=20):
            grupos = defaultdict(set)
            for d in lista_dicts:
                if d['Acción Estratégica']: grupos[d['Iniciativa']].add(f"🟦 {d['Acción Estratégica']}")
                if d['Cambio Esperado']: grupos[d['Iniciativa']].add(f"🟩 {d['Cambio Esperado']}")
                
            nombres_grupos = list(grupos.keys())
            patrones_encontrados = set()
            
            for i in range(len(nombres_grupos)):
                for j in range(i + 1, len(nombres_grupos)):
                    interseccion = frozenset(grupos[nombres_grupos[i]].intersection(grupos[nombres_grupos[j]]))
                    if len(interseccion) >= min_r and any(c.startswith("🟦") for c in interseccion) and any(c.startswith("🟩") for c in interseccion):
                        if len(interseccion) <= max_r: patrones_encontrados.add(interseccion)
                        if len(interseccion) <= 8:
                            for r in range(min_r, len(interseccion)):
                                for combo in itertools.combinations(interseccion, r):
                                    if any(c.startswith("🟦") for c in combo) and any(c.startswith("🟩") for c in combo):
                                        patrones_encontrados.add(frozenset(combo))
                                        
            resultados = {}
            for pat in patrones_encontrados:
                apariciones = [g for g, items in grupos.items() if pat.issubset(items)]
                if len(apariciones) > 1: resultados[pat] = apariciones
                    
            patrones_limpios = {}
            for pat, aps in resultados.items():
                es_redundante = False
                for pat_otro, aps_otro in resultados.items():
                    if pat != pat_otro and pat.issubset(pat_otro) and len(aps) == len(aps_otro):
                        es_redundante = True
                        break
                if not es_redundante:
                    patrones_limpios[tuple(sorted(list(pat)))] = aps
                    
            ordenados = sorted(patrones_limpios.items(), key=lambda x: (len(x[1]), len(x[0])), reverse=True)
            if not ordenados:
                st.write("No hay patrones mixtos exitosos que se repitan en múltiples Iniciativas.")
                return
            for combo, grupos_list in ordenados[:30]:
                freq = len(grupos_list)
                tamano = len(combo)
                lista_md = "\n".join([f"  {item}" for item in sorted(list(combo))])
                donde_se_vio = ", ".join(sorted(grupos_list))
                st.warning(f"**[Ecosistema de {tamano} Nodos] - Compartido con éxito por {freq} Iniciativas:**\n\n{lista_md}", icon="🌍")
                with st.expander("👁️ Ver cuáles iniciativas comparten este modelo"):
                    st.write(f"**Iniciativas:** {donde_se_vio}")

        tab_c1, tab_c2, tab_c3 = st.tabs(["📌 En una misma HISTORIA", "📌 En una misma INICIATIVA", "🌐 Ecosistema MIXTO"])
        with tab_c1:
            cc1, cc2 = st.columns(2)
            with cc1:
                st.markdown("**Acciones aplicadas simultáneamente con éxito:**")
                algoritmo_seguro(datos_exitosos, 'Historia_Cod', 'Acción Estratégica', "Historia", 2, 6)
            with cc2:
                st.markdown("**Efectos logrados simultáneamente:**")
                algoritmo_seguro(datos_exitosos, 'Historia_Cod', 'Cambio Esperado', "Historia", 2, 14)
        with tab_c2:
            cc3, cc4 = st.columns(2)
            with cc3:
                st.markdown("**Acciones exitosas movilizadas por la misma Iniciativa:**")
                algoritmo_seguro(datos_exitosos, 'Iniciativa', 'Acción Estratégica', "Iniciativa", 2, 6)
            with cc4:
                st.markdown("**Cambios logrados por la misma Iniciativa:**")
                algoritmo_seguro(datos_exitosos, 'Iniciativa', 'Cambio Esperado', "Iniciativa", 2, 14)
        with tab_c3:
            st.markdown("**Combinación Exitosa de Acciones + Cambios en la misma Iniciativa:**")
            algoritmo_mixto_veloz(datos_exitosos, 2, 20)

# ==========================================
# PÁGINA 4: REPORTE EJECUTIVO MEL
# ==========================================
elif pagina_actual == "📈 Reporte Ejecutivo MEL":
    st.markdown("### 📈 Reporte Ejecutivo MEL")
    st.caption("Visión macroestructural del Portafolio Institucional (Gráficos optimizados para exportación a PowerPoint)")
    
    df_conn = pd.DataFrame(datos_list)
    df_mel_resumen = df_catalogo.copy()
    
    if df_conn.empty: df_conn = pd.DataFrame(columns=['Iniciativa', 'Portafolio', 'Historia_Cod', 'Acción Estratégica', 'Cambio Esperado', 'Estado'])
        
    hists_por_ini = df_conn.groupby('Iniciativa')['Historia_Cod'].nunique().reset_index(name='Historias')
    conns_por_ini = df_conn.groupby('Iniciativa').size().reset_index(name='Conexiones')
    
    df_mel_resumen = df_mel_resumen.merge(hists_por_ini, on='Iniciativa', how='left')
    df_mel_resumen = df_mel_resumen.merge(conns_por_ini, on='Iniciativa', how='left')
    df_mel_resumen['Historias'] = df_mel_resumen['Historias'].fillna(0).astype(int)
    df_mel_resumen['Conexiones'] = df_mel_resumen['Conexiones'].fillna(0).astype(int)
    
    # ---------------- ANÁLISIS 1: DISTRIBUCIÓN ----------------
    st.markdown("#### 1. Distribución del Portafolio")
    tot_inis = len(df_mel_resumen)
    tot_hists = df_mel_resumen['Historias'].sum()
    
    dist_port = df_mel_resumen.groupby('Portafolio').agg(
        Iniciativas=('Iniciativa', 'nunique'),
        Historias=('Historias', 'sum'),
        Conexiones=('Conexiones', 'sum')
    ).reset_index()
    
    dist_port['% Iniciativas Num'] = (dist_port['Iniciativas'] / max(1, tot_inis)) * 100
    dist_port['% Historias Num'] = (dist_port['Historias'] / max(1, tot_hists)) * 100
    
    dist_port_view = dist_port.copy()
    dist_port_view['% Iniciativas'] = dist_port_view['% Iniciativas Num'].round(1).astype(str) + '%'
    dist_port_view['% Historias'] = dist_port_view['% Historias Num'].round(1).astype(str) + '%'
    dist_port_view = dist_port_view[['Portafolio', 'Iniciativas', '% Iniciativas', 'Historias', '% Historias', 'Conexiones']]
    
    st.dataframe(dist_port_view, use_container_width=True, hide_index=True)
    
    c_a1_1, c_a1_2, c_a1_3 = st.columns(3)
    with c_a1_1:
        fig_a1_1 = px.bar(dist_port, x='Portafolio', y='Iniciativas', title='Iniciativas por Portafolio', text_auto=True, color_discrete_sequence=['#003366'])
        fig_a1_1.update_layout(template="plotly_white", xaxis_title="")
        st.plotly_chart(fig_a1_1, use_container_width=True)
    with c_a1_2:
        fig_a1_2 = px.bar(dist_port, x='Portafolio', y='Historias', title='Historias por Portafolio', text_auto=True, color_discrete_sequence=['#FFB300'])
        fig_a1_2.update_layout(template="plotly_white", xaxis_title="")
        st.plotly_chart(fig_a1_2, use_container_width=True)
    with c_a1_3:
        fig_a1_3 = px.bar(dist_port, x='Portafolio', y='% Historias Num', title='% Historias por Portafolio', text_auto='.1f', color_discrete_sequence=['#4CAF50'])
        fig_a1_3.update_layout(template="plotly_white", xaxis_title="")
        fig_a1_3.update_traces(texttemplate='%{y:.1f}%')
        st.plotly_chart(fig_a1_3, use_container_width=True)
    st.markdown("---")

    # ---------------- ANÁLISIS 2: MADUREZ ----------------
    st.markdown("#### 2. Madurez de Cambio por Portafolio")
    if not df_conn.empty:
        madurez = df_conn.groupby(['Portafolio', 'Estado']).size().reset_index(name='Conexiones')
        totales = madurez.groupby('Portafolio')['Conexiones'].transform('sum')
        madurez['Porcentaje'] = (madurez['Conexiones'] / totales * 100).round(1)
        
        fig_madurez = px.bar(madurez, x='Portafolio', y='Conexiones', color='Estado', 
                             color_discrete_map=MEL_COLORS, text='Porcentaje',
                             title="Distribución de Estados de Conexión (Volumen y Porcentaje)")
        fig_madurez.update_traces(texttemplate='%{text}%')
        fig_madurez.update_layout(template="plotly_white", barmode='stack', legend_title="Estado de Madurez")
        st.plotly_chart(fig_madurez, use_container_width=True)
    else: st.info("No hay datos suficientes.")
    st.markdown("---")

    # ---------------- ANÁLISIS 3: INICIATIVAS LÍDERES ----------------
    st.markdown("#### 3. Iniciativas Líderes (Índice de Madurez)")
    if not df_conn.empty:
        df_conn['Es_Maduro'] = df_conn['Estado'].isin(['Negro (Ejemplo de Cambio)', 'Rojo (Señal de Buen Camino)'])
        df_conn['Es_Intencion'] = df_conn['Estado'] == 'Naranja (Intención de Cambio)'
        df_conn['Es_Reto'] = df_conn['Estado'] == 'Morado (Efecto Estancado)'
        
        madurez_ini = df_conn.groupby('Iniciativa')[['Es_Maduro', 'Es_Intencion', 'Es_Reto']].sum().reset_index()
        madurez_ini = pd.merge(df_catalogo[['Iniciativa']], madurez_ini, on='Iniciativa', how='left').fillna(0)
        
        c_l1, c_l2, c_l3 = st.columns(3)
        with c_l1:
            st.write("**TOP INICIATIVAS (Ejemplos + Señales)**")
            top1 = madurez_ini.sort_values('Es_Maduro', ascending=False).head(10)
            st.dataframe(top1[['Iniciativa', 'Es_Maduro']].rename(columns={'Es_Maduro': 'Logros'}), hide_index=True, use_container_width=True)
            fig1 = px.bar(top1, x='Es_Maduro', y='Iniciativa', orientation='h', color_discrete_sequence=['#4CAF50'], title="Mayor Evidencia de Cambio")
            fig1.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(l=0, r=0, t=30, b=0), height=300, template="plotly_white")
            st.plotly_chart(fig1, use_container_width=True)
            
        with c_l2:
            st.write("**MÁS INTENCIONES (Naranjas)**")
            top2 = madurez_ini.sort_values('Es_Intencion', ascending=False).head(10)
            st.dataframe(top2[['Iniciativa', 'Es_Intencion']].rename(columns={'Es_Intencion': 'Intenciones'}), hide_index=True, use_container_width=True)
            fig2 = px.bar(top2, x='Es_Intencion', y='Iniciativa', orientation='h', color_discrete_sequence=['#2196F3'], title="Fase de Intención")
            fig2.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(l=0, r=0, t=30, b=0), height=300, template="plotly_white")
            st.plotly_chart(fig2, use_container_width=True)
            
        with c_l3:
            st.write("**MÁS RETOS (Morados)**")
            top3 = madurez_ini.sort_values('Es_Reto', ascending=False).head(10)
            st.dataframe(top3[['Iniciativa', 'Es_Reto']].rename(columns={'Es_Reto': 'Retos'}), hide_index=True, use_container_width=True)
            fig3 = px.bar(top3, x='Es_Reto', y='Iniciativa', orientation='h', color_discrete_sequence=['#F44336'], title="Fase de Estancamiento")
            fig3.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(l=0, r=0, t=30, b=0), height=300, template="plotly_white")
            st.plotly_chart(fig3, use_container_width=True)
    st.markdown("---")

    # ---------------- ANÁLISIS 4: ROLES ESTRATÉGICOS ----------------
    st.markdown("#### 4. Roles Estratégicos (Acciones) por Portafolio")
    if not df_conn.empty:
        roles_port = pd.crosstab(df_conn['Portafolio'], df_conn['Acción Estratégica'], normalize='index') * 100
        st.write("**Tabla de Distribución Porcentual**")
        st.dataframe(roles_port.round(1).astype(str) + '%', use_container_width=True)
        
        roles_melt = roles_port.reset_index().melt(id_vars='Portafolio', var_name='Acción Estratégica', value_name='Porcentaje')
        fig_roles = px.bar(roles_melt, x='Portafolio', y='Porcentaje', color='Acción Estratégica',
                           title="Rol Predominante por Portafolio (100%)",
                           color_discrete_sequence=px.colors.qualitative.Prism, text_auto='.1f')
        fig_roles.update_layout(template="plotly_white", barmode='stack')
        st.plotly_chart(fig_roles, use_container_width=True)
    st.markdown("---")

    # ---------------- ANÁLISIS 5: CAMBIOS ESTRATÉGICOS (Puro) ----------------
    st.markdown("#### 5. Cambios Estratégicos Logrados (Solo Evidencia y Buen Camino)")
    st.caption("Filtro Estricto: Sólo analiza conexiones cuyo estado es 'Ejemplo de Cambio' o 'Señal de Buen Camino'. Omite por completo Intenciones y Efectos Estancados.")
    df_evidencia = df_conn[df_conn['Estado'].isin(['Negro (Ejemplo de Cambio)', 'Rojo (Señal de Buen Camino)'])]
    
    if not df_evidencia.empty:
        df_evidencia['Cambio_Corto'] = df_evidencia['Cambio Esperado'].map(dict_cambios)
        cam_port_pct = pd.crosstab(df_evidencia['Portafolio'], df_evidencia['Cambio_Corto'], normalize='index') * 100
        
        st.write("**Distribución Porcentual de Cambios Exitosos por Portafolio**")
        st.dataframe(cam_port_pct.round(1).astype(str) + '%', use_container_width=True)
        
        cam_melt = cam_port_pct.reset_index().melt(id_vars='Portafolio', var_name='Cambio Estratégico', value_name='Porcentaje')
        fig_cam = px.bar(cam_melt, x='Portafolio', y='Porcentaje', color='Cambio Estratégico',
                         title="Cambio Predominante por Portafolio (100% - Solo Éxitos)",
                         color_discrete_sequence=px.colors.qualitative.Bold, text_auto='.1f')
        fig_cam.update_layout(template="plotly_white", barmode='stack')
        st.plotly_chart(fig_cam, use_container_width=True)
        
        with st.expander("📚 Ver Leyenda de Cambios (C1 a C15)"):
            for k,v in dict_cambios.items(): st.markdown(f"**{v}:** {k}")
    else:
        st.info("No hay datos de evidencia (Negros/Rojos) para mostrar en este análisis.")
    st.markdown("---")
    
    # ---------------- ANÁLISIS 6: HEATMAP DE CAMBIOS ----------------
    st.markdown("#### 6. Concentración de Cambios Totales (Heatmap Máx. 15 Cambios)")
    if not df_conn.empty:
        cam_port_nunique = df_conn.groupby('Portafolio')['Cambio Esperado'].nunique().reset_index(name='Diversidad de Cambios')
        st.write("**Diversidad (Cuántos cambios distintos atiende cada Portafolio)**")
        st.dataframe(cam_port_nunique, hide_index=True)
        
        st.write("**Mapa de Calor: Concentración General**")
        df_conn['Cambio_Corto'] = df_conn['Cambio Esperado'].map(dict_cambios)
        heat_df = pd.crosstab(df_conn['Portafolio'], df_conn['Cambio_Corto'])
        
        fig_heat_mel = px.imshow(heat_df, color_continuous_scale='Blues', text_auto=True, aspect="auto")
        fig_heat_mel.update_layout(template="plotly_white", xaxis_title="Cambios Esperados (C1 a C15)", yaxis_title="Portafolio")
        st.plotly_chart(fig_heat_mel, use_container_width=True)
    st.markdown("---")

    # ---------------- ANÁLISIS 7: ALINEACIÓN ESTRATÉGICA ----------------
    st.markdown("#### 7. Alineación Estratégica por Portafolio")
    st.caption("Una historia es desalineada si apunta a 'Cambio - no en estrategia'. Una iniciativa es desalineada si tiene al menos 1 historia desalineada.")
    if not df_conn.empty:
        df_conn['No_Alineada_Hist'] = df_conn['Cambio Esperado'] == 'Cambio - no en estrategia'
        
        hist_alin_df = df_conn.groupby(['Historia_Cod', 'Portafolio'])['No_Alineada_Hist'].any().reset_index()
        hist_alin_df['Estado_Alin'] = hist_alin_df['No_Alineada_Hist'].apply(lambda x: 'Desalineada' if x else 'Alineada')
        
        ini_alin_df = df_conn.groupby('Iniciativa')['No_Alineada_Hist'].any().reset_index()
        df_aliniacion = pd.merge(df_catalogo, ini_alin_df, on='Iniciativa', how='left')
        
        # Corrección del bug de negativos: forzar a Booleano antes de procesar
        df_aliniacion['No_Alineada_Hist'] = df_aliniacion['No_Alineada_Hist'].fillna(False).astype(bool)
        df_aliniacion['Estado_Alin'] = df_aliniacion['No_Alineada_Hist'].apply(lambda x: 'Desalineada' if x else 'Alineada')
        
        res_ini = df_aliniacion.groupby(['Portafolio', 'Estado_Alin']).size().reset_index(name='Cantidad_Iniciativas')
        res_hist = hist_alin_df.groupby(['Portafolio', 'Estado_Alin']).size().reset_index(name='Cantidad_Historias')
        
        st.write("**Tabla Resumen Consolidada (Iniciativas e Historias)**")
        piv_ini = res_ini.pivot(index='Portafolio', columns='Estado_Alin', values='Cantidad_Iniciativas').fillna(0).astype(int)
        piv_ini.columns = [f'Iniciativas {c}' for c in piv_ini.columns]
        piv_hist = res_hist.pivot(index='Portafolio', columns='Estado_Alin', values='Cantidad_Historias').fillna(0).astype(int)
        piv_hist.columns = [f'Historias {c}' for c in piv_hist.columns]
        
        res_tabla = piv_ini.merge(piv_hist, on='Portafolio', how='outer').fillna(0).astype(int).reset_index()
        for expected_col in ['Iniciativas Alineada', 'Iniciativas Desalineada', 'Historias Alineada', 'Historias Desalineada']:
            if expected_col not in res_tabla.columns: res_tabla[expected_col] = 0
                
        res_tabla = res_tabla[['Portafolio', 'Iniciativas Alineada', 'Iniciativas Desalineada', 'Historias Alineada', 'Historias Desalineada']]
        res_tabla.columns = ['Portafolio', 'Iniciativas Alineadas', 'Iniciativas Desalineadas', 'Historias Alineadas', 'Historias Desalineadas']
        st.dataframe(res_tabla, use_container_width=True, hide_index=True)

        c_al1, c_al2 = st.columns(2)
        with c_al1:
            fig_al1 = px.bar(res_ini, x='Portafolio', y='Cantidad_Iniciativas', color='Estado_Alin', barmode='group',
                             color_discrete_map={'Alineada': '#4CAF50', 'Desalineada': '#D32F2F'}, text_auto=True,
                             title="Iniciativas Alineadas vs Desalineadas")
            fig_al1.update_layout(template="plotly_white")
            st.plotly_chart(fig_al1, use_container_width=True)
            
        with c_al2:
            fig_al2 = px.bar(res_hist, x='Portafolio', y='Cantidad_Historias', color='Estado_Alin', barmode='group',
                             color_discrete_map={'Alineada': '#4CAF50', 'Desalineada': '#D32F2F'}, text_auto=True,
                             title="Historias Alineadas vs Desalineadas")
            fig_al2.update_layout(template="plotly_white")
            st.plotly_chart(fig_al2, use_container_width=True)

# ==========================================
# PÁGINA 5: INDICADORES EJECUTIVOS (DIRECTIVOS)
# ==========================================
elif pagina_actual == "📊 Indicadores Ejecutivos para Dirección":
    st.markdown("### 📊 Indicadores Ejecutivos para Dirección")
    st.caption("Módulo de analítica de alto nivel diseñado para toma de decisiones ágiles.")
    
    # 0. VALIDACIÓN ESTRUCTURAL (Para auditoría del directivo)
    total_inis_universo = len(df_catalogo)
    total_cambios = len(cambios_unicos)
    
    if total_inis_universo != 44 or total_cambios > 15:
        st.error(f"⚠️ **ADVERTENCIA DE INTEGRIDAD:** El sistema detecta {total_inis_universo} Iniciativas (Esperadas: 44) y {total_cambios} Cambios (Máximo Esperado: 15). Revise la matriz de datos original.", icon="🚨")
    else:
        st.success(f"✅ **SISTEMA VALIDADO:** 44 Iniciativas cargadas | {total_cambios} Cambios detectados (Incluyendo 'No en estrategia').", icon="✅")
    st.markdown("---")

    df_conn = pd.DataFrame(datos_list)
    if df_conn.empty: df_conn = pd.DataFrame(columns=['Iniciativa', 'Portafolio', 'Historia_Cod', 'Acción Estratégica', 'Cambio Esperado', 'Estado'])

    # --- ANÁLISIS A: ALINEACIÓN ESTRATÉGICA EJECUTIVA ---
    st.markdown("#### A. Alineación Estratégica Ejecutiva")
    df_conn['No_Alineada'] = df_conn['Cambio Esperado'] == 'Cambio - no en estrategia'
    
    # Cálculos a nivel de historia
    hist_alin_df = df_conn.groupby('Historia_Cod')['No_Alineada'].any().reset_index()
    total_hists_alineadas = (~hist_alin_df['No_Alineada']).sum()
    total_hists_desalineadas = hist_alin_df['No_Alineada'].sum()
    historias_des_lista = hist_alin_df[hist_alin_df['No_Alineada']]['Historia_Cod'].tolist()
    
    # Cálculos a nivel de iniciativa cruzado con el catálogo de 44
    ini_alin_df = df_conn.groupby('Iniciativa')['No_Alineada'].any().reset_index()
    df_aliniacion_ejec = pd.merge(df_catalogo, ini_alin_df, on='Iniciativa', how='left')
    df_aliniacion_ejec['No_Alineada'] = df_aliniacion_ejec['No_Alineada'].fillna(False).astype(bool)
    
    total_inis_alineadas = (~df_aliniacion_ejec['No_Alineada']).sum()
    total_inis_desalineadas = df_aliniacion_ejec['No_Alineada'].sum()
    inis_des_lista = df_aliniacion_ejec[df_aliniacion_ejec['No_Alineada']]['Iniciativa'].tolist()

    kpi1, kpi2, kpi3, kpi4 = st.columns(4)
    kpi1.metric("Historias Alineadas", int(total_hists_alineadas))
    kpi2.metric("Historias Desalineadas", int(total_hists_desalineadas))
    kpi3.metric("Iniciativas Alineadas", int(total_inis_alineadas))
    kpi4.metric("Iniciativas Desalineadas", int(total_inis_desalineadas))
    
    col_a1, col_a2 = st.columns(2)
    with col_a1:
        st.write("**Lista de Iniciativas Desalineadas:**")
        if inis_des_lista: st.dataframe(pd.DataFrame({"Iniciativas": inis_des_lista}), use_container_width=True)
        else: st.info("Ninguna iniciativa está desalineada.")
    with col_a2:
        st.write("**Lista de Historias Desalineadas:**")
        if historias_des_lista: st.dataframe(pd.DataFrame({"Historias": historias_des_lista}), use_container_width=True)
        else: st.info("Ninguna historia está desalineada.")
    st.markdown("---")

    # --- ANÁLISIS B: CONTRIBUCIÓN A LOS CAMBIOS ESTRATÉGICOS ---
    st.markdown("#### B. Contribución a los Cambios Estratégicos")
    if not df_conn.empty:
        dist_cambios = df_conn.groupby('Cambio Esperado').size().reset_index(name='Conexiones')
        fig_donut = px.pie(dist_cambios, names='Cambio Esperado', values='Conexiones', hole=0.4, 
                           title="¿A cuáles cambios estamos aportando más? (Conexiones Totales)")
        fig_donut.update_traces(textposition='inside', textinfo='percent+label')
        fig_donut.update_layout(template="plotly_white", showlegend=False, height=500)
        st.plotly_chart(fig_donut, use_container_width=True)
    st.markdown("---")

    # --- ANÁLISIS C: TOP 3 INICIATIVAS POR PORTAFOLIO (Evidencia) ---
    st.markdown("#### C. Top 3 Iniciativas por Portafolio (Mayor Evidencia)")
    st.caption("Basado exclusivamente en conexiones con estado: Negro (Ejemplo) + Rojo (Buen Camino).")
    df_exito = df_conn[df_conn['Estado'].isin(['Negro (Ejemplo de Cambio)', 'Rojo (Señal de Buen Camino)'])]
    
    if not df_exito.empty:
        top_inis_port = df_exito.groupby(['Portafolio', 'Iniciativa']).size().reset_index(name='Conexiones_Evidenciadas')
        top_inis_port = top_inis_port.sort_values(['Portafolio', 'Conexiones_Evidenciadas'], ascending=[True, False])
        top3_inis = top_inis_port.groupby('Portafolio').head(3).reset_index(drop=True)
        
        c_c1, c_c2 = st.columns([1, 2])
        with c_c1:
            st.dataframe(top3_inis, hide_index=True, use_container_width=True)
        with c_c2:
            fig_top_inis = px.bar(top3_inis, x='Conexiones_Evidenciadas', y='Iniciativa', color='Portafolio', 
                                  orientation='h', text_auto=True, title="Iniciativas Líderes por Portafolio")
            fig_top_inis.update_layout(template="plotly_white", yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig_top_inis, use_container_width=True)
    else: st.info("No hay evidencia suficiente (Negros o Rojos).")
    st.markdown("---")

    # Variables globales para mapeo de D, E, F, G
    portafolios_unicos = [p for p in df_catalogo['Portafolio'].unique() if p.strip() != '']
    cambios_unicos_list = list(set(d['Cambio Esperado'] for d in datos_list))
    roles_unicos_list = list(set(d['Acción Estratégica'] for d in datos_list))

    # --- ANÁLISIS D & E: CAMBIOS MÁS Y MENOS MOVILIZADOS ---
    st.markdown("#### D & E. Movilización de Cambios por Portafolio")
    st.caption("D: Cambios con más conexiones. | E: Oportunidades de mejora (Cambios con 0 o menor presencia).")
    
    idx_cambios = pd.MultiIndex.from_product([portafolios_unicos, cambios_unicos_list], names=['Portafolio', 'Cambio Esperado'])
    df_base_cambios = pd.DataFrame(index=idx_cambios).reset_index()
    conteos_cambios = df_conn.groupby(['Portafolio', 'Cambio Esperado']).size().reset_index(name='Frecuencia')
    df_full_cambios = pd.merge(df_base_cambios, conteos_cambios, on=['Portafolio', 'Cambio Esperado'], how='left').fillna(0)
    
    # Calcular Porcentajes internos del portafolio
    totales_cambios_port = df_full_cambios.groupby('Portafolio')['Frecuencia'].transform('sum')
    df_full_cambios['Porcentaje'] = ((df_full_cambios['Frecuencia'] / totales_cambios_port) * 100).fillna(0).round(1).astype(str) + '%'
    df_full_cambios['Cambio_Corto'] = df_full_cambios['Cambio Esperado'].map(dict_cambios)

    top3_mas_cambios = df_full_cambios.sort_values(['Portafolio', 'Frecuencia'], ascending=[True, False]).groupby('Portafolio').head(3).reset_index(drop=True)
    top3_menos_cambios = df_full_cambios.sort_values(['Portafolio', 'Frecuencia'], ascending=[True, True]).groupby('Portafolio').head(3).reset_index(drop=True)

    c_de1, c_de2 = st.columns(2)
    with c_de1:
        st.write("**Top 3 MÁS Movilizados**")
        st.dataframe(top3_mas_cambios[['Portafolio', 'Cambio_Corto', 'Frecuencia', 'Porcentaje']], hide_index=True, use_container_width=True)
    with c_de2:
        st.write("**Top 3 MENOS Movilizados (Oportunidad de Mejora)**")
        st.dataframe(top3_menos_cambios[['Portafolio', 'Cambio_Corto', 'Frecuencia']], hide_index=True, use_container_width=True)
    with st.expander("📚 Leyenda de Cambios (C1 a C15)"):
        for k,v in dict_cambios.items(): st.markdown(f"**{v}:** {k}")
    st.markdown("---")

    # --- ANÁLISIS F & G: ROLES MÁS Y MENOS EJERCIDOS ---
    st.markdown("#### F & G. Capacidades Estratégicas (Roles) por Portafolio")
    st.caption("F: Roles más utilizados. | G: Capacidades menos ejercidas en el sistema.")
    
    idx_roles = pd.MultiIndex.from_product([portafolios_unicos, roles_unicos_list], names=['Portafolio', 'Acción Estratégica'])
    df_base_roles = pd.DataFrame(index=idx_roles).reset_index()
    conteos_roles = df_conn.groupby(['Portafolio', 'Acción Estratégica']).size().reset_index(name='Frecuencia')
    df_full_roles = pd.merge(df_base_roles, conteos_roles, on=['Portafolio', 'Acción Estratégica'], how='left').fillna(0)
    
    totales_roles_port = df_full_roles.groupby('Portafolio')['Frecuencia'].transform('sum')
    df_full_roles['Porcentaje'] = ((df_full_roles['Frecuencia'] / totales_roles_port) * 100).fillna(0).round(1).astype(str) + '%'
    df_full_roles['Accion_Corta'] = df_full_roles['Acción Estratégica'].map(dict_acciones)

    top3_mas_roles = df_full_roles.sort_values(['Portafolio', 'Frecuencia'], ascending=[True, False]).groupby('Portafolio').head(3).reset_index(drop=True)
    top3_menos_roles = df_full_roles.sort_values(['Portafolio', 'Frecuencia'], ascending=[True, True]).groupby('Portafolio').head(3).reset_index(drop=True)

    c_fg1, c_fg2 = st.columns(2)
    with c_fg1:
        st.write("**Top 3 Roles MÁS Ejercidos**")
        st.dataframe(top3_mas_roles[['Portafolio', 'Accion_Corta', 'Frecuencia', 'Porcentaje']], hide_index=True, use_container_width=True)
    with c_fg2:
        st.write("**Top 3 Roles MENOS Ejercidos**")
        st.dataframe(top3_menos_roles[['Portafolio', 'Accion_Corta', 'Frecuencia']], hide_index=True, use_container_width=True)
    with st.expander("📚 Leyenda de Roles (A1 a A6)"):
        for k,v in dict_acciones.items(): st.markdown(f"**{v}:** {k}")
